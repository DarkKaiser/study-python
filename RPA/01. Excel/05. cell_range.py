from random import *
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])

for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # B컬럼('영어')의 모든 데이터만 가져온다.
for cell in col_B:
    print(cell.value)
    
col_Range = ws["B:C"] # B컬럼('영어')부터 C컬럼('수학')의 모든 데이터를 가져온다.
for cols in col_Range:
    for cell in cols:
        print(cell.value)
        
row_title = ws[1]
for cell in row_title:
    print(cell.value)
    
row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for row in row_range:
    for cell in row:
        print(cell.coordinate, end=":")
        
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=":")
        
        print(cell.value, end=" ")
    print()
    
# 전체 rows
print(tuple(ws.rows))

# 전체 columns
print(tuple(ws.columns))

for row in ws.iter_rows():  # 전체 row
    print(row[2].value)

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value)

wb.save("sample.xlsx")
wb.close()

from openpyxl import load_workbook

wb1 = load_workbook("sample_formula.xlsx")
ws1 = wb1.active

# 수식 그대로 가져오고 있음
for row in ws1.values:
    for cell in row:
        print(cell)

wb1.close()

#########################################

wb2 = load_workbook("sample_formula.xlsx", data_only=True)
ws2 = wb2.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은(아직 수식의 계산이 되지 않은) 상태의 데이터는 None이라고 표시
# 엑셀 파일을 열었다가 저장한 후(수식을 계산), 다시 실행해보면 제대로 된 데이터가 출력된다.
for row in ws2.values:
    for cell in row:
        print(cell)

wb2.close()

from random import *
from openpyxl import Workbook

wb = Workbook()         # 새 워크북 생성

ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1

print(ws["A1"])         # A1 셀의 정보를 출력
print(ws["A1"].value)   # A1 셀의 '값'을 출력
print(ws["A10"].value)  # 값이 없을 땐 'None'을 출력

# ws["A1"].value와 동일
print(ws.cell(row=1, column=1).value)

# ws["C1"] = 10과 동일
c = ws.cell(row=1, column=3, value=10)
print(c.value)  # ws["C1"] 셀의 '값'을 출력

# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자

wb.save("sample.xlsx")
wb.close()

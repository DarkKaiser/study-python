from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2부터 D2까지 셀 병합
ws["B2"] = "Merged Cell"

wb.save("sample_cell_merge.xlsx")
wb.close()

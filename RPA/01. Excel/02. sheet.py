from openpyxl import Workbook

wb = Workbook()         # 새 워크북 생성

ws = wb.create_sheet()                  # 새로운 시트 기본 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet(title="YourSheet")    # 주어진 이름으로 시트 생성

ws2 = wb.create_sheet("NewSheet", index=2)  # 2번째 인덱스에 시트 생성

new_ws = wb["NewSheet"] # 딕셔너리 형태로 시트에 접근

# 전체 시트의 이름 목록 출력
print(wb.sheetnames)

# 시트 복사
new_ws["A1"] = "Test"   # A1 셀에 Test라는 값을 입력
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
wb.close()
